# excel_utils.py — работа с Excel

from openpyxl import Workbook, load_workbook
import os
from config import LOG_FILE, RECEIPTS_DIR

def init_excel():
    # Создаем необходимые директории
    os.makedirs(RECEIPTS_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
    
    if not os.path.exists(LOG_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        headers = [
            "Дата", "Имя пользователя", "User ID", "Страна", 
            "Сервис", "Курс (руб)", "Сумма (USD)", "Сумма (руб)", "Статус"
        ]
        ws.append(headers)
        wb.save(LOG_FILE)

def save_order(user_data, status="Ожидание"):
    wb = load_workbook(LOG_FILE)
    ws = wb.active
    row = [
        user_data["date"],
        user_data["username"],
        user_data["user_id"],
        user_data["country"],
        user_data["service"],
        user_data["rate"],
        user_data["amount_usd"],
        user_data["amount_rub"],
        status
    ]
    ws.append(row)
    wb.save(LOG_FILE)

def get_stats():
    if not os.path.exists(LOG_FILE):
        return 0
    
    wb = load_workbook(LOG_FILE)
    ws = wb.active
    total = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            if isinstance(row[7], (int, float)):  # Сумма в рублях
                total += row[7]
        except:
            continue
    
    return total